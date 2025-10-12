import json
import re
from pathlib import Path
from typing import Any, List

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter, range_boundaries
from PyPDF2 import PdfReader

from excel_copilot.tools.excel_tools import translate_range_with_references


def _load_pdf_text(path: Path, collapse_spaces: bool) -> str:
    reader = PdfReader(str(path))
    chunks: List[str] = []
    for page in reader.pages:
        text = page.extract_text()
        if text:
            chunks.append(text)
    joined = "\n".join(chunks)
    if collapse_spaces:
        joined = re.sub(r"\s+", "", joined)
    else:
        joined = re.sub(r"\s+", " ", joined)
    return joined


def _extract_japanese_sentences(pdf_path: Path, count: int) -> List[str]:
    text = _load_pdf_text(pdf_path, collapse_spaces=True)
    sentences = [segment for segment in text.split("。") if len(segment) > 15]
    return [sentence + "。" for sentence in sentences[:count]]


def _extract_english_sentences(pdf_path: Path, count: int) -> List[str]:
    text = _load_pdf_text(pdf_path, collapse_spaces=False)
    sentences = re.split(r"(?<=[.!?])\s+", text)
    selected: List[str] = []
    for sentence in sentences:
        cleaned = sentence.strip()
        if len(cleaned) > 60:
            selected.append(cleaned)
        if len(selected) >= count:
            break
    return selected


class StubWorkbookActions:
    def __init__(self, workbook_path: Path, sheet_name: str) -> None:
        self.workbook_path = workbook_path
        self.sheet_name = sheet_name
        self.progress_log: List[str] = []
        self.write_calls: List[dict[str, Any]] = []
        self.book = type("Book", (), {"fullname": str(workbook_path), "full_name": str(workbook_path)})
        self._wb = load_workbook(workbook_path)

    def _get_ws(self):
        return self._wb[self.sheet_name]

    def read_range(self, cell_range: str, sheet_name: str | None = None):
        ws = self._get_ws()
        min_col, min_row, max_col, max_row = range_boundaries(cell_range)
        data: List[List[Any]] = []
        for row in ws.iter_rows(
            min_row=min_row,
            max_row=max_row,
            min_col=min_col,
            max_col=max_col,
            values_only=True,
        ):
            data.append(list(row))
        return data

    def write_range(self, cell_range: str, values: List[List[Any]], sheet_name: str | None = None):
        ws = self._get_ws()
        min_col, min_row, max_col, max_row = range_boundaries(cell_range)
        assert len(values) == max_row - min_row + 1
        for r_offset, row_values in enumerate(values):
            assert len(row_values) == max_col - min_col + 1
            for c_offset, value in enumerate(row_values):
                ws.cell(row=min_row + r_offset, column=min_col + c_offset).value = value
        self._wb.save(self.workbook_path)
        self.write_calls.append({"range": cell_range, "values": values})
        start = f"{get_column_letter(min_col)}{min_row}"
        end = f"{get_column_letter(max_col)}{max_row}"
        return f"Wrote range {start}:{end}"

    def log_progress(self, message: str) -> None:
        self.progress_log.append(message)


class StubBrowserManager:
    def __init__(self, jp_sentences: List[str], en_sentences: List[str]) -> None:
        self.jp_sentences = jp_sentences
        self.en_sentences = en_sentences

    def ask(self, prompt: str, stop_event=None) -> str:
        prompt = prompt.strip()
        if "キーフレーズ" in prompt:
            payload = [{"key_phrases": ["企業理念", "研究開発", "財務状況", "市場環境", "サステナビリティ", "経営方針"]}]
            return json.dumps(payload, ensure_ascii=False)

        if "原文（日本語）に対応する参照資料から関連する文章を抽出してください" in prompt:
            payload = [{"source_sentences": self.jp_sentences}]
            return json.dumps(payload, ensure_ascii=False)

        if "原文参照文と" in prompt and "\"pairs\"" in prompt:
            payload = [
                {
                    "pairs": [
                        {"source_sentence": jp, "target_sentence": en}
                        for jp, en in zip(self.jp_sentences, self.en_sentences)
                    ]
                }
            ]
            return json.dumps(payload, ensure_ascii=False)

        if "translated_text" in prompt:
            payload = [
                {
                    "translated_text": "Based on the reference passages, the consolidated outlook was prepared with careful risk assessment.",
                    "process_notes_jp": "参照資料の該当箇所を確認し、用語と数値の整合性を維持して翻訳しました。",
                    "reference_pairs": [
                        {"source_sentence": jp, "target_sentence": en}
                        for jp, en in zip(self.jp_sentences, self.en_sentences)
                    ],
                }
            ]
            return json.dumps(payload, ensure_ascii=False)

        raise AssertionError(f"Unhandled prompt: {prompt[:120]}")


def test_reference_sentences_exist_in_pdf(tmp_path: Path):
    jp_pdf = Path("/Users/ralleti/Downloads/f_repo250624.pdf")
    en_pdf = Path("/Users/ralleti/Downloads/f_repo250627 (1).pdf")
    jp_sentences = _extract_japanese_sentences(jp_pdf, count=2)
    en_sentences = _extract_english_sentences(en_pdf, count=2)
    assert len(jp_sentences) == 2 and len(en_sentences) == 2

    wb_path = tmp_path / "autotest.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet"
    ws["A1"] = "テスト原文です。"
    wb.save(wb_path)
    wb.close()

    actions = StubWorkbookActions(wb_path, sheet_name="Sheet")
    browser_manager = StubBrowserManager(jp_sentences=jp_sentences, en_sentences=en_sentences)

    message = translate_range_with_references(
        actions=actions,
        browser_manager=browser_manager,
        cell_range="Sheet!A1",
        target_language="English",
        sheet_name="Sheet",
        source_reference_urls=[str(jp_pdf)],
        target_reference_urls=[str(en_pdf)],
        translation_output_range="Sheet!B1:I1",
        overwrite_source=False,
    )

    assert "Wrote range B1:I1" in message

    loaded = load_workbook(wb_path)
    ws_loaded = loaded["Sheet"]
    translation_value = ws_loaded["B1"].value
    process_notes_value = ws_loaded["C1"].value
    pairs = [ws_loaded[get_column_letter(col) + "1"].value for col in range(4, 10)]
    loaded.close()

    assert translation_value is not None
    assert process_notes_value is not None

    pdf_jp_text = _load_pdf_text(jp_pdf, collapse_spaces=True)
    pdf_en_text = _load_pdf_text(en_pdf, collapse_spaces=False)

    for index, sentence in enumerate(jp_sentences, start=1):
        assert sentence in pdf_jp_text
        expected_pair = f"{sentence}\n---\n{en_sentences[index - 1]}"
        assert pairs[index - 1] == expected_pair
        assert en_sentences[index - 1] in pdf_en_text

    for remaining in pairs[len(jp_sentences):]:
        assert remaining in (None, "")
