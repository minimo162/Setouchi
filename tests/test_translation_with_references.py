import json
from pathlib import Path
from typing import Any, List, Optional

import pytest
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter, range_boundaries

from excel_copilot.tools.excel_tools import translate_range_with_references


class _DummyBook:
    def __init__(self, full_path: Path) -> None:
        # translate_range_contents で fullname/full_name を参照するためのダミー
        self.fullname = str(full_path)
        self.full_name = str(full_path)


class StubActions:
    """translate_range_with_references が必要とする最小限の ExcelActions 互換クラス"""

    def __init__(self, workbook_path: Path, sheet_name: str) -> None:
        self.workbook_path = workbook_path
        self.sheet_name = sheet_name
        self.book = _DummyBook(workbook_path)
        self._wb = load_workbook(workbook_path)
        self.progress_log: List[str] = []
        self.write_calls: List[dict[str, Any]] = []

    def _get_ws(self, sheet_name: Optional[str]):
        name = sheet_name or self.sheet_name
        return self._wb[name]

    def read_range(self, cell_range: str, sheet_name: Optional[str] = None):
        ws = self._get_ws(sheet_name)
        min_col, min_row, max_col, max_row = range_boundaries(cell_range)
        data: List[List[Any]] = []
        for row in ws.iter_rows(
            min_row=min_row,
            max_row=max_row,
            min_col=min_col,
            max_col=max_col,
            values_only=True,
        ):
            data.append(list(row))
        return data

    def write_range(self, cell_range: str, values: List[List[Any]], sheet_name: Optional[str] = None):
        ws = self._get_ws(sheet_name)
        min_col, min_row, max_col, max_row = range_boundaries(cell_range)
        expected_rows = max_row - min_row + 1
        expected_cols = max_col - min_col + 1
        assert len(values) == expected_rows, "行数が一致しません"
        for row_offset, row_values in enumerate(values):
            assert len(row_values) == expected_cols, "列数が一致しません"
            for col_offset, cell_value in enumerate(row_values):
                ws.cell(row=min_row + row_offset, column=min_col + col_offset).value = cell_value
        self._wb.save(self.workbook_path)
        self.write_calls.append(
            {
                "range": cell_range,
                "sheet": sheet_name or self.sheet_name,
                "values": values,
            }
        )
        start = f"{get_column_letter(min_col)}{min_row}"
        end = f"{get_column_letter(max_col)}{max_row}"
        return f"Wrote range {start}:{end}"

    def log_progress(self, message: str) -> None:
        self.progress_log.append(message)


class StubBrowserManager:
    """Copilot への問い合わせをモックし、テスト用の決め打ちレスポンスを返す"""

    def __init__(self) -> None:
        self.calls: List[str] = []

    def ask(self, prompt: str, stop_event=None) -> str:
        prompt = prompt.strip()
        if "キーフレーズを日本語で6個作成してください" in prompt:
            self.calls.append("key_phrases")
            payload = [
                {
                    "key_phrases": [
                        "関税政策",
                        "業績予想",
                        "米国政府",
                        "慎重判断",
                        "現時点情報",
                        "連結業績",
                    ]
                }
            ]
            return json.dumps(payload, ensure_ascii=False)

        if "原文（日本語）に対応する参照資料から関連する文章を抽出してください" in prompt:
            self.calls.append("source_sentences")
            payload = [
                {
                    "source_sentences": [
                        "現時点で入手可能な情報や予測を踏まえ、連結業績予想を算定しました。",
                        "米国政府の関税政策動向を慎重に見極めながら方針を決定した。",
                    ]
                }
            ]
            return json.dumps(payload, ensure_ascii=False)

        if "原文参照文と" in prompt and "\"pairs\"" in prompt:
            self.calls.append("pairs")
            payload = [
                {
                    "pairs": [
                        {
                            "source_sentence": "現時点で入手可能な情報や予測を踏まえ、連結業績予想を算定しました。",
                            "target_sentence": "Based on the information and forecasts currently available, we calculated the consolidated earnings outlook.",
                        },
                        {
                            "source_sentence": "米国政府の関税政策動向を慎重に見極めながら方針を決定した。",
                            "target_sentence": "We determined our approach while carefully assessing the trajectory of U.S. government tariff policies.",
                        },
                    ]
                }
            ]
            return json.dumps(payload, ensure_ascii=False)

        # 最終翻訳プロンプト
        if "translated_text" in prompt:
            self.calls.append("translation")
            payload = [
                {
                    "translated_text": (
                        "We had kept the consolidated forecast for fiscal year 2026 undecided while "
                        "closely monitoring U.S. government tariff policies, but based on the "
                        "information and projections currently available, we calculated the outlook "
                        "as stated above and are now announcing it."
                    ),
                    "process_notes_jp": (
                        "参照文が示す慎重な関税政策への注視と、入手可能な情報に基づく予測算定を反映して訳語を決定。"
                        "ビジネス文書らしい文調に整えた。"
                    ),
                    "reference_pairs": [
                        {
                            "source_sentence": "現時点で入手可能な情報や予測を踏まえ、連結業績予想を算定しました。",
                            "target_sentence": "Based on the information and forecasts currently available, we calculated the consolidated earnings outlook.",
                        },
                        {
                            "source_sentence": "米国政府の関税政策動向を慎重に見極めながら方針を決定した。",
                            "target_sentence": "We determined our approach while carefully assessing the trajectory of U.S. government tariff policies.",
                        },
                    ],
                }
            ]
            return json.dumps(payload, ensure_ascii=False)

        self.calls.append("unhandled")
        raise ValueError(f"Unhandled prompt: {prompt[:80]}...")


@pytest.fixture()
def workbook_path(tmp_path: Path) -> Path:
    """テスト用に A1 に原文を設定した Excel ファイルを生成する"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet"
    ws["A1"] = (
        "2026年３月期の通期連結業績予想につきましては、米国政府の関税政策の動向などを慎重に見極めるため、未\n"
        "定としておりましたが、現時点で入手可能な情報や予測を踏まえ、上記の通り連結業績予想を算定しましたので公\n"
        "表するものであります。"
    )
    path = tmp_path / "test.xlsx"
    wb.save(path)
    wb.close()
    return path


def test_translate_range_with_references_pipeline(workbook_path: Path):
    actions = StubActions(workbook_path, sheet_name="Sheet")
    browser_manager = StubBrowserManager()

    try:
        message = translate_range_with_references(
            actions=actions,
            browser_manager=browser_manager,
            cell_range="Sheet!A1",
            target_language="English",
            sheet_name="Sheet",
            source_reference_urls=[
                "https://example.com/source_reference.pdf",
            ],
            target_reference_urls=[
                "https://example.com/target_reference.pdf",
            ],
            translation_output_range="Sheet!B1:I1",
            overwrite_source=False,
        )
    except Exception:
        # デバッグ用にどのプロンプトが呼ばれたかを表示して再送出
        print("Browser calls:", browser_manager.calls)
        raise

    # 期待するログ／書き込みが行われたか確認
    assert "Wrote range B1:I1" in message
    assert any("キーフレーズ生成" in log for log in actions.progress_log)
    assert any("日本語参照文章抽出" in log for log in actions.progress_log)
    assert any("対になる英語参照文抽出" in log for log in actions.progress_log)

    wb = load_workbook(workbook_path)
    ws = wb["Sheet"]
    translation = ws["B1"].value
    process_notes = ws["C1"].value
    pair_1 = ws["D1"].value
    pair_2 = ws["E1"].value
    pair_3 = ws["F1"].value
    pair_4 = ws["G1"].value
    pair_5 = ws["H1"].value
    pair_6 = ws["I1"].value
    wb.close()

    assert translation is not None and "consolidated forecast" in translation
    assert process_notes is not None and "参照文" in process_notes
    assert pair_1 == (
        "現時点で入手可能な情報や予測を踏まえ、連結業績予想を算定しました。\n"
        "---\n"
        "Based on the information and forecasts currently available, we calculated the consolidated earnings outlook."
    )
    assert pair_2 == (
        "米国政府の関税政策動向を慎重に見極めながら方針を決定した。\n"
        "---\n"
        "We determined our approach while carefully assessing the trajectory of U.S. government tariff policies."
    )
    assert pair_3 in (None, "")
    assert pair_4 in (None, "")
    assert pair_5 in (None, "")
    assert pair_6 in (None, "")
